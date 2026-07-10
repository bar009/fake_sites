"""Write the run outputs: results.xlsx, report.html, results.json.

Every row is one (brand x search result) as a plain dict - see COLUMNS.
"""

import html
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("brand", "Brand"),
    ("topic", "Topic"),
    ("rank", "Rank"),
    ("url", "URL"),
    ("final_url", "Final URL"),
    ("page_title", "Page Title"),
    ("domain", "Domain"),
    ("domain_created", "Domain Created"),
    ("domain_age_days", "Domain Age (days)"),
    ("registrar", "Registrar"),
    ("country", "Country"),
    ("flags", "Suspicion Flags"),
    ("note", "Note"),
    ("screenshot", "Screenshot"),
    ("error", "Error"),
]


def write_json(rows: list[dict], path: Path) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def write_xlsx(rows: list[dict], path: Path, extra_columns: tuple = ()) -> None:
    columns = COLUMNS + list(extra_columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333355")
    suspicious_fill = PatternFill("solid", fgColor="FFD7D7")

    for col, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        suspicious = bool(row.get("flags"))
        for col, (key, _) in enumerate(columns, start=1):
            value = row.get(key, "")
            if key == "flags":
                value = "; ".join(value) if value else ""
            cell = ws.cell(row=row_idx, column=col, value=value)
            if key in ("url", "final_url") and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
            if suspicious:
                cell.fill = suspicious_fill

    widths = {"brand": 18, "topic": 14, "url": 45, "final_url": 45, "page_title": 30,
              "domain": 28, "registrar": 24, "flags": 32, "note": 30, "screenshot": 34, "error": 40}
    for col, (key, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(key, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def _slug(text: str) -> str:
    return re.sub(r"[^\w-]+", "-", text).strip("-").lower() or "x"


def write_html(rows: list[dict], path: Path, run_name: str) -> None:
    by_brand: dict[str, list[dict]] = {}
    for row in rows:
        by_brand.setdefault(row["brand"], []).append(row)

    parts = [f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<title>Fake Shop Report - {html.escape(run_name)}</title>
<style>
  body {{ font-family: "Segoe UI", Arial, sans-serif; background: #f4f4f7; margin: 0; padding: 24px; color: #222; }}
  h1 {{ font-size: 22px; }}
  .brand {{ background: #fff; border-radius: 10px; padding: 16px 20px; margin-bottom: 24px; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  .brand h2 {{ margin: 0 0 12px; font-size: 18px; }}
  .cards {{ display: flex; flex-wrap: wrap; gap: 16px; }}
  .card {{ width: 320px; border: 1px solid #ddd; border-radius: 8px; padding: 10px; background: #fafafa; }}
  .card.suspicious {{ border-color: #d33; background: #fff3f3; }}
  .card img {{ width: 100%; height: 200px; object-fit: cover; object-position: top; border: 1px solid #ccc; border-radius: 4px; }}
  .card .meta {{ font-size: 12.5px; direction: ltr; text-align: left; word-break: break-all; margin-top: 8px; }}
  .flag {{ display: inline-block; background: #d33; color: #fff; border-radius: 4px; padding: 1px 7px; font-size: 12px; margin: 2px 2px 0 0; direction: ltr; }}
  .err {{ color: #b00; font-size: 12.5px; direction: ltr; text-align: left; }}
  a {{ color: #0645ad; }}
  .none {{ color: #777; font-size: 14px; }}
  .summary {{ background: #fff3f3; border: 1px solid #d33; border-radius: 10px; padding: 14px 20px; margin-bottom: 24px; }}
  .summary h2 {{ margin: 0 0 8px; font-size: 16px; color: #b00; }}
  .summary li {{ margin: 3px 0; }}
  .topic-head {{ font-size: 15px; color: #555; margin: 26px 0 10px; border-bottom: 1px solid #ccc; padding-bottom: 4px; }}
</style>
</head>
<body>
<h1>דו"ח אתרים חשודים — {html.escape(run_name)}</h1>
<p>סה"כ מותגים: {len(by_brand)} | סה"כ תוצאות: {len(rows)}</p>
"""]

    flagged = [r for r in rows if r.get("flags")]
    if flagged:
        parts.append(f'<div class="summary"><h2>🚩 {len(flagged)} אתרים חשודים</h2><ul>')
        for r in flagged:
            age = r.get("domain_age_days")
            age_txt = f"{age} days" if age is not None else "?"
            parts.append(
                f'<li><a href="#brand-{_slug(r["brand"])}">{html.escape(r["brand"])}</a> — '
                f'<span dir="ltr">{html.escape(r.get("domain", ""))} ({age_txt})</span></li>')
        parts.append('</ul></div>')

    last_topic = None
    for brand, brand_rows in by_brand.items():
        topic = brand_rows[0].get("topic") or ""
        if topic != last_topic:
            parts.append(f'<div class="topic-head" dir="ltr">{html.escape(topic or "(no topic)")}</div>')
            last_topic = topic
        topic_html = f' <small>({html.escape(topic)})</small>' if topic else ""
        parts.append(f'<div class="brand" id="brand-{_slug(brand)}"><h2>{html.escape(brand)}{topic_html}</h2>')
        real_rows = [r for r in brand_rows if r.get("url")]
        if not real_rows:
            parts.append('<p class="none">לא נמצאו תוצאות חיפוש</p></div>')
            continue
        parts.append('<div class="cards">')
        for row in real_rows:
            suspicious = " suspicious" if row.get("flags") else ""
            parts.append(f'<div class="card{suspicious}">')
            shot = row.get("screenshot")
            if shot:
                # master report rows carry a screenshot_href pointing into their run folder
                shot_href = row.get("screenshot_href") or f"screenshots/{shot}"
                parts.append(f'<a href="{html.escape(shot_href)}" target="_blank">'
                             f'<img src="{html.escape(shot_href)}" alt="screenshot"></a>')
            url = row.get("url", "")
            age = row.get("domain_age_days")
            age_txt = f"{age} days old" if age is not None else "age unknown"
            registrar = row.get("registrar") or "registrar unknown"
            parts.append(f'<div class="meta">#{row.get("rank")} '
                         f'<a href="{html.escape(url)}" target="_blank" rel="noopener noreferrer">{html.escape(url)}</a><br>'
                         f'{html.escape(row.get("domain", ""))} — {html.escape(age_txt)} — {html.escape(registrar)}</div>')
            for flag in row.get("flags") or []:
                parts.append(f'<span class="flag">{html.escape(flag)}</span>')
            if row.get("note"):
                parts.append(f'<div class="meta" style="color:#886">{html.escape(row["note"])}</div>')
            if row.get("error"):
                parts.append(f'<div class="err">{html.escape(row["error"])}</div>')
            parts.append('</div>')
        parts.append('</div></div>')

    parts.append("</body></html>")
    path.write_text("\n".join(parts), encoding="utf-8")
